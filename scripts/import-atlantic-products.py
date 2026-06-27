import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = Path(r"C:\Users\Ning Sun\Desktop\推送_三引擎标准复核版.xlsx")
PRODUCTS_PATH = ROOT / "data" / "jinqiao-products.json"
REPORT_PATH = ROOT / "docs" / "atlantic-import-report.md"

CATEGORY_MAP = {
    "碳钢焊条": ("carbon-steel-electrodes", "碳钢焊条"),
    "实芯气保及氩弧焊丝": ("solid-wires", "实芯气保及氩弧焊丝"),
    "药芯气保焊丝": ("flux-cored-wires", "药芯气保焊丝"),
    "不锈钢焊材": ("stainless-materials", "不锈钢焊材"),
    "碳钢埋弧焊丝焊剂": ("submerged-arc", "碳钢埋弧焊丝焊剂"),
    "埋弧焊丝焊剂": ("submerged-arc", "碳钢埋弧焊丝焊剂"),
    "铝焊丝": ("aluminum-wires", "铝焊丝"),
    "特种焊材": ("special-materials", "特种焊材"),
}


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("–", "-").replace("—", "-").replace("－", "-")
    lines = [" ".join(line.split()) for line in text.split("\n")]
    return "\n".join(line for line in lines if line).strip()


def clean_inline(value):
    return " ".join(clean_text(value).split())


def model_from_raw(raw):
    model = clean_inline(raw)
    model = re.sub(r"^(焊丝|焊条|焊剂)", "", model).strip()
    return model


def slugify_model(model):
    slug = model.lower()
    slug = slug.replace("·", "-").replace("/", "-").replace("\\", "-")
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = re.sub(r"-+", "-", slug).strip("-")
    return slug or "unknown"


def split_standards(*values):
    standards = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        for part in re.split(r"[\n；;]+", text):
            item = clean_inline(part)
            if item and item not in standards:
                standards.append(item)
    return standards


def parse_key_values(value):
    text = clean_text(value)
    if not text:
        return []
    rows = []
    for part in re.split(r"[；;\n]+", text):
        item = clean_inline(part)
        if not item:
            continue
        match = re.match(r"^([^:：]+)[:：]\s*(.+)$", item)
        if match:
            name = clean_inline(match.group(1))
            val = clean_inline(match.group(2))
        else:
            name = item
            val = ""
        if name:
            rows.append({"name": name, "value": val})
    return rows


def first_sentence(*values):
    text = clean_inline(next((value for value in values if clean_inline(value)), ""))
    if not text:
        return ""
    match = re.search(r"[。.!！?？]", text)
    if match:
        return text[: match.end()]
    return text[:120]


def read_source_pages(workbook):
    if "标准交叉检查" not in workbook.sheetnames:
        return {}
    ws = workbook["标准交叉检查"]
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    row_index = header.index("行号")
    source_index = header.index("来源页")
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[row_index] and row[source_index]:
            result[int(row[row_index])] = clean_inline(row[source_index])
    return result


def build_products():
    workbook = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    source_pages = read_source_pages(workbook)
    sheet = workbook["Sheet1"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    raw_items = []
    skipped_blank = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        if clean_inline(data.get("品牌")) != "大西洋":
            continue
        raw_model = clean_inline(data.get("牌号"))
        if not raw_model:
            continue
        model = model_from_raw(raw_model)
        if not model:
            skipped_blank.append(row_number)
            continue
        category_slug, category_name = CATEGORY_MAP.get(clean_inline(data.get("分类")), ("special-materials", "特种焊材"))
        standards = split_standards(data.get("国标型号"), data.get("美标型号"), data.get("国际标型号"))
        introduction = clean_text(data.get("介绍"))
        application = clean_text(data.get("特性"))
        summary_detail = first_sentence(introduction, application)
        summary = f"大西洋 {model}，{category_name}。"
        if summary_detail:
            summary += summary_detail
        source_page = source_pages.get(row_number)
        source = "推送_三引擎标准复核版.xlsx"
        if source_page:
            source += f"；来源页：{source_page}"
        product = {
            "slug": f"atlantic-{slugify_model(model)}",
            "manufacturer": "大西洋",
            "categorySlug": category_slug,
            "categoryName": category_name,
            "model": model,
            "name": f"大西洋 {model} {category_name}",
            "standard": " / ".join(standards),
            "standards": standards,
            "summary": summary,
            "introduction": introduction,
            "applications": [application] if application else [],
            "composition": parse_key_values(data.get("成分")),
            "depositedMetal": parse_key_values(data.get("熔覆金属数据")),
            "dimensions": [],
            "certifications": [],
            "notes": "以大西洋最新质保书、产品样本及项目技术条件为准；关键工程请结合 WPS/PQR 复核。",
            "source": source,
            "_sourceRow": row_number,
        }
        raw_items.append(product)

    by_slug = {}
    duplicate_rows = []
    for product in raw_items:
        slug = product["slug"]
        if slug in by_slug:
            duplicate_rows.append((slug, by_slug[slug]["_sourceRow"], product["_sourceRow"]))
            current_score = data_score(by_slug[slug])
            next_score = data_score(product)
            if next_score > current_score:
                by_slug[slug] = product
        else:
            by_slug[slug] = product

    products = []
    for product in by_slug.values():
        product.pop("_sourceRow", None)
        products.append(product)
    products.sort(key=lambda item: (item["categorySlug"], item["model"]))
    return products, skipped_blank, duplicate_rows


def data_score(product):
    return sum(
        [
            bool(product.get("standards")),
            bool(product.get("introduction")),
            bool(product.get("applications")),
            bool(product.get("composition")),
            bool(product.get("depositedMetal")),
        ]
    )


def main():
    atlantic_products, skipped_blank, duplicate_rows = build_products()
    existing = json.loads(PRODUCTS_PATH.read_text(encoding="utf-8-sig"))
    retained = [item for item in existing if item.get("manufacturer") != "大西洋"]
    merged = retained + atlantic_products
    PRODUCTS_PATH.write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    category_counts = Counter(item["categoryName"] for item in atlantic_products)
    missing = {
        "介绍": sum(not item.get("introduction") for item in atlantic_products),
        "适用场景": sum(not item.get("applications") for item in atlantic_products),
        "执行标准": sum(not item.get("standards") for item in atlantic_products),
        "成分": sum(not item.get("composition") for item in atlantic_products),
        "熔覆金属": sum(not item.get("depositedMetal") for item in atlantic_products),
    }
    report = [
        "# 大西洋产品导入报告",
        "",
        f"- 来源文件：`{EXCEL_PATH.name}`",
        f"- 导入大西洋产品：{len(atlantic_products)} 个",
        f"- 跳过空型号行：{len(skipped_blank)} 行（{', '.join(map(str, skipped_blank)) or '无'}）",
        f"- 重复 slug 合并：{len(duplicate_rows)} 组",
        "",
        "## 分类数量",
        "",
    ]
    for name, count in sorted(category_counts.items()):
        report.append(f"- {name}: {count}")
    report.extend(["", "## 缺失字段统计", ""])
    for name, count in missing.items():
        report.append(f"- {name}: {count}")
    if duplicate_rows:
        report.extend(["", "## 重复合并记录", ""])
        for slug, first_row, duplicate_row in duplicate_rows:
            report.append(f"- `{slug}`: 保留信息更完整记录，来源行 {first_row} / {duplicate_row}")
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text("\n".join(report) + "\n", encoding="utf-8")
    print(json.dumps({"imported": len(atlantic_products), "skippedBlank": skipped_blank, "duplicates": duplicate_rows, "missing": missing}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
